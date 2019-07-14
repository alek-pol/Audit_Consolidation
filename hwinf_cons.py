import os
import glob
import lxml
import math
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill

src_path = 'c:\Development\Audit\src\*'
dst_path = 'c:\Development\Audit\\'
dst_file = 'Test1.xlsx'

min_win10_build = 16299  # 1709
sncp = {
    'Raw': {'SheetName': 'Raw', 'CurRow': 1, 'CurCol': 0},
    'List': {'SheetName': 'List', 'CurRow': 0, 'CurCol': 0},
    'Consolidation': {'SheetName': 'Consolidation', 'CurRow': 0, 'CurCol': 0}
}
sncp_lock = list(sncp.keys())
list_col = []

out_data_raw = {
    'Organithation': {
        'name': 'Наименование организации', 'b_style': 'Main_left', 'width': 30, 'value': ''},
    'Workplace': {
        'name': 'Рабочее место', 'b_style': 'Main_left', 'width': 16, 'value': ''},
    'Computer Name:': {
        'name': 'Название АРМ', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Case Type:': {
        'name': 'Тип корпуса', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Status': {
        'name': 'Статус', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Operating System:': {
        'name': 'Операционная система', 'b_style': 'Main_left', 'width': 20, 'value': ''},
    'OS need update': {
        'name': 'Необходимо обновить ОС', 'b_style': 'Main_center', 'width': 11, 'value': ''},
    'Number Of Processor Cores:': {
        'name': 'Кол-во ядер', 'b_style': 'Main_center', 'width': 8, 'value': ''},
    'Number Of Logical Processors:': {
        'name': 'Кол-во логич-х проц-в', 'b_style': 'Main_center', 'width': 9, 'value': ''},
    'CPU Brand Name:': {
        'name': 'Процессор', 'b_style': 'Main_center', 'width': 20, 'value': ''},
    'CPU Platform:': {
        'name': 'Сокет', 'b_style': 'Main_center', 'width': 10, 'value': ''},
    'L3 Cache:': {
        'name': 'L3 - Кэш', 'b_style': 'Main_center', 'width': 10, 'value': ''},
    'Motherboard Model:': {
        'name': 'Модель мат.платы', 'b_style': 'Main_center', 'width': 15, 'value': ''},
    'Data': {
        'name': 'Год произв-ва', 'b_style': 'Main_center', 'width': 7, 'value': ''},
    'Total Memory Size:': {
        'name': 'Опер. памяти всего', 'b_style': 'Main_center', 'width': 10, 'value': ''},
    'Total Memory Size [MB]:': {
        'name': 'Опер. памяти всего (Mb)', 'b_style': 'Main_center', 'width': 10, 'value': ''},
    'Memory upgrade': {
        'name': 'Необходимо увеличить память', 'b_style': 'Main_center', 'width': 11, 'value': ''},
    'Maximum Supported Memory Clock:': {
        'name': 'Максимально поддерживаемая скороость', 'b_style': 'Main_center', 'width': 10, 'value': ''},
    'Maximum Memory Size per Channel:': {
        'name': 'Максимум памяти на канал', 'b_style': 'Main_center', 'width': 10, 'value': ''},
    'Memory Type:': {
        'name': 'Тип памяти', 'b_style': 'Main_center', 'width': 10, 'value': ''},
    'Memory sale': {
        'name': 'Наличие памяти в продаже', 'b_style': 'Main_center', 'width': 10, 'value': ''},


    'Module Type:': {
        'name': 'Вид модуля', 'b_style': 'Main_center', 'width': 10, 'value': ''},
    'Total Memory Count': {
        'name': 'Кол-во модулей', 'b_style': 'Main_center', 'width': 10, 'value': ''},


    'Module Density:0': {
        'name': 'Модуль 1 объем', 'b_style': 'Main_center', 'width': 10, 'value': ''},
    'Memory Speed:0': {
        'name': 'Модуль 1 скорость', 'b_style': 'Main_center', 'width': 10, 'value': ''},
    'Module Manufacturer:0': {
        'name': 'Модуль 1 производитель', 'b_style': 'Main_left', 'width': 13, 'value': ''},
    'Module Part Number:0': {
        'name': 'Модуль 1 PartNumber', 'b_style': 'Main_left', 'width': 10, 'value': ''},

    'Module Density:1': {
        'name': 'Модуль 2 объем', 'b_style': 'Main_center', 'width': 10, 'value': ''},
    'Memory Speed:1': {
        'name': 'Модуль 2 скорость', 'b_style': 'Main_center', 'width': 10, 'value': ''},
    'Module Manufacturer:1': {
        'name': 'Модуль 2 производитель', 'b_style': 'Main_left', 'width': 13, 'value': ''},
    'Module Part Number:1': {
        'name': 'Модуль 2 PartNumber', 'b_style': 'Main_left', 'width': 10, 'value': ''},

    'Module Density:2': {
        'name': 'Модуль 3 объем', 'b_style': 'Main_center', 'width': 10, 'value': ''},
    'Memory Speed:2': {
        'name': 'Модуль 3 скорость', 'b_style': 'Main_center', 'width': 10, 'value': ''},
    'Module Manufacturer:2': {
        'name': 'Модуль 3 производитель', 'b_style': 'Main_left', 'width': 13, 'value': ''},
    'Module Part Number:2': {
        'name': 'Модуль 3 PartNumber', 'b_style': 'Main_left', 'width': 10, 'value': ''},

    'Module Density:3': {
        'name': 'Модуль 4 объем', 'b_style': 'Main_center', 'width': 10, 'value': ''},
    'Memory Speed:3': {
        'name': 'Модуль 4 скорость', 'b_style': 'Main_center', 'width': 10, 'value': ''},
    'Module Manufacturer:3': {
        'name': 'Модуль 4 производитель', 'b_style': 'Main_left', 'width': 13, 'value': ''},
    'Module Part Number:3': {
        'name': 'Модуль 4 PartNumber', 'b_style': 'Main_left', 'width': 10, 'value': ''},

    'Installed SSD': {
        'name': 'Установлен SSD', 'b_style': 'Main_center', 'width': 10, 'value': ''},

    'Drive Controller:0': {
        'name': 'Привод 1 контроллер', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Drive Model:0': {
        'name': 'Привод 1 модель', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Drive Capacity [MB]:0': {
        'name': 'Привод 1 объем', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Media Rotation Rate:0': {
        'name': 'Привод 1 скорость вращения', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Nominal Form Factor:0': {
        'name': 'Привод 1 Form Factor', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Cache Buffer Size:0': {
        'name': 'Привод 1 объем буфера', 'b_style': 'Main_left', 'width': 10, 'value': ''},

    'Drive Controller:1': {
        'name': 'Привод 2 контроллер', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Drive Model:1': {
        'name': 'Привод 2 модель', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Drive Capacity [MB]:1': {
        'name': 'Привод 2 объем', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Media Rotation Rate:1': {
        'name': 'Привод 2 скорость вращения', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Nominal Form Factor:1': {
        'name': 'Привод 2 Form Factor', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Cache Buffer Size:1': {
        'name': 'Привод 2 объем буфера', 'b_style': 'Main_left', 'width': 10, 'value': ''},

    'Drive Controller:2': {
        'name': 'Привод 3 контроллер', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Drive Model:2': {
        'name': 'Привод 3 модель', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Drive Capacity [MB]:2': {
        'name': 'Привод 3 объем', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Media Rotation Rate:2': {
        'name': 'Привод 3 скорость вращения', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Nominal Form Factor:2': {
        'name': 'Привод 3 Form Factor', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Cache Buffer Size:2': {
        'name': 'Привод 3 объем буфера', 'b_style': 'Main_left', 'width': 10, 'value': ''},

    'Diagonal': {
        'name': 'Диагональ монитора', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Monitor Name:': {
            'name': 'Модель', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Monitor Name (Manuf):': {
            'name': 'Модельный ряд', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Serial Number:': {
            'name': 'Серийный номер', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Date Of Manufacture:': {
            'name': 'Дата производства', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Max. Vertical Size:': {
            'name': 'Высота', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Max. Horizontal Size:': {
            'name': 'Ширина', 'b_style': 'Main_left', 'width': 10, 'value': ''}
}


out_data_org = {
    'Workplace': {
        'name': 'Рабочее место', 'b_style': 'Main_left', 'width': 15, 'value': ''},
    'Status': {
        'name': 'Статус', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'CPU Brand Name:': {
        'name': 'Процессор', 'b_style': 'Main_center', 'width': 15, 'value': ''},
    'Number Of Processor Cores:': {
        'name': 'Кол-во ядер', 'b_style': 'Main_center', 'width': 5, 'value': ''},


    'CPU Platform:': {
        'name': 'Сокет', 'b_style': 'Main_center', 'width': 10, 'value': ''},
    'L3 Cache:': {
        'name': 'L3 - Кэш', 'b_style': 'Main_center', 'width': 10, 'value': ''},

    'Data': {
        'name': 'Год произв-ва', 'b_style': 'Main_center', 'width': 7, 'value': ''},
    'Total Memory Size [MB]:': {
        'name': 'Опер. памяти всего (Mb)', 'b_style': 'Main_center', 'width': 10, 'value': ''},
    'Memory upgrade': {
        'name': 'Необходимо увеличить память', 'b_style': 'Main_center', 'width': 10, 'value': ''},
    'Memory Type:': {
        'name': 'Тип памяти', 'b_style': 'Main_center', 'width': 10, 'value': ''},


    'Module Type:': {
        'name': 'Вид модуля', 'b_style': 'Main_center', 'width': 10, 'value': ''},


    'Installed SSD': {
        'name': 'Установлен SSD', 'b_style': 'Main_center', 'width': 10, 'value': ''},
    'Drive Capacity [MB]:0': {
        'name': 'Привод 1 объем', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Drive Capacity [MB]:1': {
        'name': 'Привод 2 объем', 'b_style': 'Main_left', 'width': 10, 'value': ''},


    'Diagonal': {
        'name': 'Диагональ монитора', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Monitor Name:': {
            'name': 'Модель', 'b_style': 'Main_left', 'width': 10, 'value': ''},
    'Date Of Manufacture:': {
            'name': 'Дата производства', 'b_style': 'Main_left', 'width': 10, 'value': ''},

    'Case Type:': {
        'name': 'Тип корпуса', 'b_style': 'Main_left', 'width': 10, 'value': ''},


    'Operating System:': {
        'name': 'Операционная система', 'b_style': 'Main_left', 'width': 20, 'value': ''},
    'OS need update': {
        'name': 'Необходимо обновить ОС', 'b_style': 'Main_center', 'width': 11, 'value': ''},


}



mem_cost = {
    '200.0 MHz (PC3200)': {1: 1000},
    '266.7 MHz (PC2-4200)': {2: 2050},

    '533.3 MHz (DDR3-1066 / PC3-8500)': {4: 1800, 8: 2500},
    '666.7 MHz (DDR3-1333 / PC3-10600)': {4: 2900, 8: 3600},
    '800.0 MHz (DDR3-1600 / PC3-12800)': {4: 1500, 8: 3000},
    '933.3 MHz (DDR3-1866 / PC3-14900)': {4: 2200, 8: 3750},

    '1066.1 MHz (DDR4-2132 / PC4-17000)': {4: 1900, 8: 2050},
    '1067.2 MHz (DDR4-2134 / PC4-17000)': {4: 1900, 8: 2050},
    '1200.5 MHz (DDR4-2400 / PC4-19200)': {4: 1400, 8: 5700},
    '1333.3 MHz (DDR4-2666 / PC4-21300)': {4: 1800, 8: 2500}
}

def gen_list_col():
    l_col = []
    for f_ch in ['', 'A', 'B']:
        for n_ch in range(26):
            l_col.append((f_ch + chr(65 + n_ch)))
    return l_col


def gen_cr(list_n):
    cur_cr = list_col[sncp[list_n]['CurCol']] + str(sncp[list_n]['CurRow'])
    sncp[list_n]['CurCol'] += 1
    return cur_cr


def add_style(workbook, new_style, **kwargs):
    if new_style not in workbook.named_styles:
        n_style = NamedStyle(name=new_style)
        for key, value in kwargs.items():
            setattr(n_style, key, value)
        workbook.add_named_style(n_style)


def init_style():
    b_side = Side(style='thin', color='000000')

    add_style(new_wb, 'Topic_main',
              font=Font(name='Calibri', size=12, bold=True),
              alignment=Alignment(vertical='center'),
              )

    add_style(new_wb, 'Topic_tab',
              font=Font(name='Calibri', size=9, bold=True),
              alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
              border=Border(left=b_side, top=b_side, right=b_side, bottom=b_side)
              )

    add_style(new_wb, 'Main_center',
              font=Font(name='Calibri', size=9),
              alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
              border=Border(left=b_side, top=b_side, right=b_side, bottom=b_side)
              )

    add_style(new_wb, 'Main_center_red',
              font=Font(name='Calibri', size=9),
              alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
              border=Border(left=b_side, top=b_side, right=b_side, bottom=b_side),
              fill=PatternFill(fgColor="FFA07A", patternType='solid')
              )

    add_style(new_wb, 'Main_center_yellow',
              font=Font(name='Calibri', size=9),
              alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
              border=Border(left=b_side, top=b_side, right=b_side, bottom=b_side),
              fill=PatternFill(fgColor="FFE4B5", patternType='solid')
              )

    add_style(new_wb, 'Main_center_green',
              font=Font(name='Calibri', size=9),
              alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
              border=Border(left=b_side, top=b_side, right=b_side, bottom=b_side),
              fill=PatternFill(fgColor="98FB98", patternType='solid')
              )

    add_style(new_wb, 'Main_left',
              font=Font(name='Calibri', size=9),
              alignment=Alignment(vertical='center', wrap_text=True),
              border=Border(left=b_side, top=b_side, right=b_side, bottom=b_side)
              )

    add_style(new_wb, 'Main_hyperlink',
              font=Font(name='Calibri', size=9, color='00008B'),
              alignment=Alignment(vertical='center', wrap_text=True),
              border=Border(left=b_side, top=b_side, right=b_side, bottom=b_side)
              )


def make_main_topic(sheet_n, col_row, value_n, style_n):
    sheet_n[col_row] = value_n
    sheet_n[col_row].style = style_n
    sheet_n.row_dimensions[1].height = 25


def make_tab_topic(sheet_n, col_row, value_n, style_n, dimension_n):
    sheet_n[col_row] = value_n
    sheet_n[col_row].style = style_n
    sheet_n.column_dimensions[col_row[0]].width = dimension_n


def make_tab_body(sheet_n, col_row, value_n, style_n):
    sheet_n[col_row] = value_n
    sheet_n[col_row].style = style_n


def make_list():
    make_main_topic(new_wb['List'], 'A1', 'Список листов', 'Topic_main')
    make_tab_topic(new_wb['List'], 'A3', '№ Листа', 'Topic_tab', 6)
    make_tab_topic(new_wb['List'], 'B3', 'Наименование организации', 'Topic_tab', 80)
    for key in sncp.keys():
        sncp['List']['CurRow'] += 1
        if key not in sncp_lock:
            make_tab_body(new_wb['List'], 'A%s' % sncp['List']['CurRow'], sncp[key]['SheetName'], 'Main_center')
            hyperlink ='=HYPERLINK("#%s!A1", "%s")' % (sncp[key]['SheetName'], key)
            make_tab_body(new_wb['List'], 'B%s' % sncp['List']['CurRow'], hyperlink, 'Main_hyperlink')


def make_raw_topic():
    make_main_topic(new_wb['Raw'], 'A1', 'Общий свод данных', 'Topic_main')
    sncp['Raw']['CurRow'] = 3
    sncp['Raw']['CurCol'] = 0
    for k, v in out_data_raw.items():
        make_tab_topic(new_wb['Raw'], gen_cr('Raw'), v['name'], 'Topic_tab', v['width'])
    sncp['Raw']['CurRow'] += 1


def out_raw_body():
    sncp['Raw']['CurCol'] = 0
    if out_data_raw['Organithation']['value'] != '':
        for k, v in out_data_raw.items():
            make_tab_body(new_wb['Raw'], gen_cr('Raw'), v['value'], v['b_style'])
            v['value'] = ''
        sncp['Raw']['CurRow'] += 1


def out_org_body(list_n, org_name):
    sncp[org_name]['CurCol'] = 0
    for k, v in out_data_org.items():
        make_tab_body(new_wb[list_n], gen_cr(org_name), v['value'], v['b_style'])
        v['value'] = ''
    sncp[org_name]['CurRow'] += 1

def make_org_topic(list_n, org_name):

    make_main_topic(new_wb[list_n], 'A1', org_name, 'Topic_main')
    sncp[org_name]['CurRow'] = 3
    sncp[org_name]['CurCol'] = 0
    for k, v in out_data_org.items():
        make_tab_topic(new_wb[list_n], gen_cr(org_name), v['name'], 'Topic_tab', v['width'])
    sncp[org_name]['CurRow'] = 4


def scan_hwi_htm(src_hwi_path):
    list_path = src_hwi_path.split('\\')
    llp = len(list_path)
    hwi_htm = open(src_hwi_path, 'r')
    soup = BeautifulSoup(hwi_htm, 'lxml')  # Parse the HTML as a string
    tables = soup.find_all('table')
    len_tables = len(tables)

    org = list_path[llp - 3]
    workplace = list_path[llp - 2]

    info_base = {
        'Computer Name:': '',
        'Current User Name:': '',
        'Manufacturer:': '',
        'Case Type:': ''
    }
    info_os = {
        'Operating System:': '',
        'OS Version': '',
        'OS Category': '',
        'OS Build': 0,
        'OS Type': '',
        'OS need update': 'Да',
        'Update style': 'Main_center_red'
    }
    info_date = {
        'value': '',
        'age': '',
        'style': 'Main_center'
    }
    info_cpu = {
        'Number Of Processor Cores:': '',
        'Number Of Logical Processors:': '',
        'CPU Brand Name:': '',
        'CPU Code Name:': '',
        'CPU Technology:': '',
        'CPU Platform:': '',
        'L3 Cache:': '',
        'Maximum Memory Size per Channel:': ''
    }
    info_mb = {
        "Motherboard Model:": "",
        "BIOS Date:": ""
    }
    info_memory = {
        'Total Memory Size:': '',
        'Total Memory Size [MB]:': '',
        'Maximum Supported Memory Clock:': '',
        'Current Memory Clock:': '',
        'Current Timing (tCAS-tRCD-tRP-tRAS):': '',
        'Memory Runs At:': '',
        'Memory Type:': '',
        'Module Type:': '',
        'Total Memory Count': 0,

        'Memory upgrade': '',
        'Memory upgrade style': 'Main_center',
        'Memory sale': 'Да',
        'Memory sale style': 'Main_center'
        }
    info_module = {
        'Module Size:': ['', '', '', ''],
        'Module Density:': ['', '', '', ''],
        'Memory Speed:': ['', '', '', ''],
        'Module Manufacturer:': ['', '', '', ''],
        'Module Part Number:': ['', '', '', '']
    }
    info_drives = {
        'Drive Controller:': ['', '', '', ''],
        'Drive Model:': ['', '', '', ''],
        'Drive Capacity [MB]:': ['', '', '', ''],
        'Media Rotation Rate:': ['', '', '', ''],
        'Nominal Form Factor:': ['', '', '', ''],
        'Cache Buffer Size:': ['', '', '', ''],
        'TRIM Command:': ['', '', '', '']
    }
    info_monitor = {
        'Monitor Name:': '',
        'Monitor Name (Manuf):': '',
        'Serial Number:': '',
        'Date Of Manufacture:': '',
        'Max. Vertical Size:': '',
        'Max. Horizontal Size:': '',
        'Diagonal': ''
    }


    def make_headers_list(tables_n):
        headers = []
        for t in range(len(tables_n)):
            tmp = None
            for col in tables_n[t].find_all("td", class_="dt"):
                tmp = col.get_text()
            if tmp:
                headers.append(tmp)
            else:
                headers.append('')
        return headers

    def scan_value(table_n, f_param):
        flag_p = False
        current_p = ''
        for row in table_n.find_all('tr'):
            for column in row.find_all('td'):
                tmp_s = column.get_text()
                if tmp_s in f_param and flag_p == False:
                    current_p = tmp_s
                    flag_p = True
                elif flag_p:
                    flag_p = False
                    f_param[current_p] = tmp_s.rstrip().lstrip()

    def check_date(date_m):
        date_m = int(date_m.split('/')[2])
        if int(date_m) < 100:
            date_m += 2000
        info_date['value'] = date_m
        if date_m <= 2010:
            info_date['age'] = 'old'
            info_date['style'] = 'Main_center_red'
        elif date_m >= 2011 and date_m <= 2013:
            info_date['age'] = 'middle'
            info_date['style'] = 'Main_center_yellow'
        else:
            info_date['age'] = 'new'
            info_date['style'] = 'Main_center_green'

    def check_os(os_n):
        if 'Microsoft' in os_n:
            build = os_n[os_n.index('Build') + 1].split('.')[0]
            if build:
                info_os['OS Build'] = int(build)
            if '(x64)' in os_n:
                info_os['OS Type'] = 'x64'
            if os_n[2] in ['10', '8.1', '8', '7', 'Vista', 'XP', '2000']:
                info_os['OS Version'] = os_n[2]
            elif os_n[2] == 'Server':
                info_os['OS Version'] = os_n[2] + ' ' + os_n[3]

            if info_os['OS Version'] == '10' and info_os['OS Build'] >= min_win10_build:
                info_os['OS need update'] = 'Нет'
                info_os['Update style'] = 'Main_center_green'

    def check_memory():
        info_memory['Total Memory Size [MB]:'] = int(info_memory['Total Memory Size [MB]:'])
        if info_memory['Total Memory Size [MB]:'] < 4096:
            info_memory['Memory upgrade'] = 'Да'
            info_memory['Memory upgrade style'] = 'Main_center_red'
        else:
            info_memory['Memory upgrade'] = 'Нет'

    def check_module():
        index_module = []
        index_memdev = []
        module_tmp = {
            'Module Size:': '',
            'Module Density:': '',
            'Memory Type:': '',
            'Module Type:': '',
            'Memory Speed:': '',
            'Module Manufacturer:': '',
            'Module Part Number:': ''
            }
        memdev_tmp0 = {
            'Device Size:': '',  # Module Density:
            'Device Type:': '',  # Memory Type:
            'Device Form Factor:': '',   # Module Type:
            'Memory Speed:': '',  # Memory Speed:
            'Manufacturer:': '',  # Module Manufacturer:
            'Part Number:': ''  # Module Part Number:
        }

        def wr_type(im_name, im_value):
            if im_value != '' and info_memory[im_name] == '':
                info_memory[im_name] = im_value


        for t in range(header_list.index('Memory'), len_tables-50):
            if header_list[t].startswith('Row:'):
                index_module.append(t)
        if len(index_module) != 0:
            for im in range(len(index_module)):
                scan_value(tables[index_module[im] + 1], module_tmp)
                if module_tmp['Module Size:'] != '':
                    if module_tmp['Module Density:'] == '':
                        ms = module_tmp['Module Size:'].split(' ')
                        if ms[1] == 'GBytes':
                            module_tmp['Module Density:'] = int(ms[0]) * 1024
                        else:
                            module_tmp['Module Density:'] = int(ms[0])
                    else:
                        module_tmp['Module Density:'] = module_tmp['Module Density:'].split(' ')[0]
                    for mkey in info_module.keys():
                        info_module[mkey][im] = module_tmp[mkey]
                        module_tmp[mkey] = ''
                    wr_type('Memory Type:', module_tmp['Memory Type:'])
                    wr_type('Module Type:', module_tmp['Module Type:'])
                else:
                    info_memory['Total Memory Count'] -= 1
            info_memory['Total Memory Count'] = len(index_module)
        else:
            s_start = header_list.index('Memory Devices')
            for t in range(s_start + 1, s_start + 15):
                if header_list[t] == 'Memory Device':
                    index_memdev.append(t)
            for im in range(len(index_memdev)):
                memdev_tmp = dict(memdev_tmp0)
                scan_value(tables[index_memdev[im] + 1], memdev_tmp)
                if memdev_tmp['Device Size:'] != '0 MBytes':
                    info_module['Module Density:'][im] = memdev_tmp['Device Size:']
                    info_module['Memory Speed:'][im] = memdev_tmp['Memory Speed:']
                    info_module['Module Manufacturer:'][im] = memdev_tmp['Manufacturer:']
                    info_module['Module Part Number:'][im] = memdev_tmp['Part Number:']
                    wr_type('Memory Type:', memdev_tmp['Device Type:'])
                    wr_type('Module Type:', memdev_tmp['Device Form Factor:'])
                    info_memory['Total Memory Count'] += 1
        if info_memory['Memory Type:'].startswith('DDR2'):
            info_memory['Memory sale'] = 'Нет'
            info_memory['Memory sale style'] = 'Main_center_red'

    def check_drive():
        drives_tmp0 = {
            'Drive Controller:': '',
            'Drive Model:': '',
            'Drive Capacity:': '',
            'Drive Capacity [MB]:': '',
            'Media Rotation Rate:': '',
            'Cache Buffer Size:': '',
            'TRIM Command:': '',
            'Nominal Form Factor:': '',
            'Drive Type:': '',
            'Device Type:': ''
        }
        i_drive = 0
        i_ssd = 'Нет'
        start = header_list.index('Drives')
        stop = 0
        index_drives = []
        stop = header_list.index('Audio')
        if stop == 0:
            stop = header_list.index('Network')
        for id in range(start + 1, stop):
            index_drives.append(id)
        dvd = ['DVD+R DL', 'BD-ROM']
        for dev in index_drives:
            drives_tmp = dict(drives_tmp0)
            scan_value(tables[dev], drives_tmp)
            if drives_tmp['Drive Model:'] != '' \
                    and drives_tmp['Drive Type:'] not in dvd \
                    and drives_tmp['Device Type:'] not in dvd:
                if drives_tmp['Media Rotation Rate:'] == 'SSD Drive (Non-rotating)':
                    i_ssd = 'Да'
                for mkey in info_drives.keys():
                    info_drives[mkey][i_drive] = drives_tmp[mkey]
                i_drive += 1
            drives_tmp = {}
        return i_ssd, i_drive

    def check_monitor():
        scan_value(tables[header_list.index('Monitor') + 3], info_monitor)
        try:
            v = int(" ".join(filter(lambda s: s.isnumeric(), info_monitor["Max. Vertical Size:"].split())))
            h = int(" ".join(filter(lambda s: s.isnumeric(), info_monitor["Max. Horizontal Size:"].split())))
            info_monitor['Diagonal'] = round((math.sqrt(v * v + h * h)) / 2.54)
        except:
            info_monitor['Diagonal'] = ""

        if info_monitor['Date Of Manufacture:'] != '':
            date = info_monitor['Date Of Manufacture:'].split(' ')
            year = date[len(date)-1]
            info_monitor['Date Of Manufacture:'] = year






    # ====================
    header_list = make_headers_list(tables)
    print(workplace)
    # ------------------------------------------------------------------------------------
    # Base info
    scan_value(tables[3], info_base)
    scan_value(tables[3], info_os)
    try:
        index_base = header_list.index('System Enclosure')
        scan_value(tables[index_base + 1], info_base)
    except:
        print()

    # Operating System info
    check_os(info_os["Operating System:"].split(' '))

    # CPU info
    index_cpu = header_list.index('Central Processor(s)')
    scan_value(tables[index_cpu + 1], info_cpu)
    scan_value(tables[index_cpu + 3], info_cpu)

    # Motherboard info
    index_mb = header_list.index('Motherboard')
    scan_value(tables[index_mb + 1], info_mb)
    check_date(info_mb["BIOS Date:"])

    # Memory Device
    index_memory = header_list.index('Memory')
    scan_value(tables[index_memory + 1], info_memory)
    check_memory()
    check_module()

    # Drive info
    info_ssd, drives_count = check_drive()

    # Monitor info
    check_monitor()
    #




    #if date_status == 1:
    status_t = ['На списание', 'Минимальные', 'Средние', 'Высокие']
    #else:
     #   status = ['', '']



    # entry Raw list
    out_data_raw['Organithation']['value'] = org
    out_data_raw['Workplace']['value'] = workplace
    out_data_raw['Computer Name:']['value'] = info_base['Computer Name:']
    out_data_raw['Case Type:']['value'] = info_base['Case Type:']
    #out_data_raw['Status']['value'] = status[0]

    out_data_raw['Operating System:']['value'] = info_os['Operating System:']
    out_data_raw['OS need update']['value'] = info_os['OS need update']
    out_data_raw['OS need update']['b_style'] = info_os['Update style']

    out_data_raw['Number Of Processor Cores:']['value'] = info_cpu['Number Of Processor Cores:']
    out_data_raw['Number Of Logical Processors:']['value'] = info_cpu['Number Of Logical Processors:']
    out_data_raw['CPU Brand Name:']['value'] = info_cpu['CPU Brand Name:']
    out_data_raw['CPU Platform:']['value'] = info_cpu['CPU Platform:']
    out_data_raw['L3 Cache:']['value'] = info_cpu['L3 Cache:']

    out_data_raw['Motherboard Model:']['value'] = info_mb['Motherboard Model:']
    out_data_raw['Data']['value'] = info_date['value']
    out_data_raw['Data']['b_style'] = info_date['style']

    out_data_raw['Total Memory Size:']['value'] = info_memory['Total Memory Size:']
    out_data_raw['Total Memory Size [MB]:']['value'] = info_memory['Total Memory Size [MB]:']
    out_data_raw['Memory upgrade']['value'] = info_memory['Memory upgrade']
    out_data_raw['Memory upgrade']['b_style'] = info_memory['Memory upgrade style']

    out_data_raw['Maximum Supported Memory Clock:']['value'] = info_memory['Maximum Supported Memory Clock:']
    out_data_raw['Maximum Memory Size per Channel:']['value'] = info_cpu['Maximum Memory Size per Channel:']
    out_data_raw['Memory Type:']['value'] = info_memory['Memory Type:']
    out_data_raw['Module Type:']['value'] = info_memory['Module Type:']
    out_data_raw['Memory sale']['value'] = info_memory['Memory sale']
    out_data_raw['Memory sale']['b_style'] = info_memory['Memory sale style']

    out_data_raw['Total Memory Count']['value'] = info_memory['Total Memory Count']
    for i in range(info_memory['Total Memory Count']):
        out_data_raw['Module Density:%i' % i]['value'] = info_module['Module Density:'][i]
        out_data_raw['Memory Speed:%s' % i]['value'] = info_module['Memory Speed:'][i]
        out_data_raw['Module Manufacturer:%s' % i]['value'] = info_module['Module Manufacturer:'][i]
        out_data_raw['Module Part Number:%s' % i]['value'] = info_module['Module Part Number:'][i]

    out_data_raw['Installed SSD']['value'] = info_ssd
    for i in range(drives_count):
        if i == 3:
            break
        out_data_raw['Drive Model:%i' % i]['value'] = info_drives['Drive Model:'][i]
        out_data_raw['Drive Controller:%i' % i]['value'] = info_drives['Drive Controller:'][i]
        out_data_raw['Drive Capacity [MB]:%i' % i]['value'] = info_drives['Drive Capacity [MB]:'][i]
        out_data_raw['Media Rotation Rate:%i' % i]['value'] = info_drives['Media Rotation Rate:'][i]
        out_data_raw['Nominal Form Factor:%i' % i]['value'] = info_drives['Nominal Form Factor:'][i]
        out_data_raw['Cache Buffer Size:%i' % i]['value'] = info_drives['Cache Buffer Size:'][i]

    out_data_raw['Monitor Name:']['value'] = info_monitor['Monitor Name:']
    out_data_raw['Monitor Name (Manuf):']['value'] = info_monitor['Monitor Name (Manuf):']
    out_data_raw['Serial Number:']['value'] = info_monitor['Serial Number:']
    out_data_raw['Date Of Manufacture:']['value'] = info_monitor['Date Of Manufacture:']
    out_data_raw['Max. Vertical Size:']['value'] = info_monitor['Max. Vertical Size:']
    out_data_raw['Max. Horizontal Size:']['value'] = info_monitor['Max. Horizontal Size:']
    out_data_raw['Diagonal']['value'] = info_monitor['Diagonal']

    out_raw_body()
    # entry Org list
    out_data_org['Workplace']['value'] = workplace
    out_data_org['Status']['value'] = ''
    out_data_org['Number Of Processor Cores:']['value'] = info_cpu['Number Of Processor Cores:']
    out_data_org['CPU Brand Name:']['value'] = info_cpu['CPU Brand Name:']
    out_data_org['CPU Platform:']['value'] = info_cpu['CPU Platform:']
    out_data_org['Data']['value'] = info_date['value']
    out_data_org['Total Memory Size [MB]:']['value'] = info_memory['Total Memory Size [MB]:']
    out_data_org['Memory upgrade']['value'] = info_memory['Memory upgrade']

    out_data_org['Memory Type:']['value'] = info_memory['Memory Type:']
    out_data_org['Module Type:']['value'] =  info_memory['Module Type:']
    out_data_org['Installed SSD']['value'] = info_ssd
    out_data_org['Drive Capacity [MB]:0']['value'] = info_drives['Drive Capacity [MB]:'][0]
    out_data_org['Drive Capacity [MB]:1']['value'] = info_drives['Drive Capacity [MB]:'][1]
    out_data_org['Diagonal']['value'] = info_monitor['Diagonal']
    out_data_org['Monitor Name:']['value'] = info_monitor['Monitor Name:']
    out_data_org['Date Of Manufacture:']['value'] = info_monitor['Date Of Manufacture:']
    out_data_org['Operating System:']['value'] = info_os['Operating System:']
    out_data_org['OS need update']['value'] = info_os['OS need update']
    out_data_org['Case Type:']['value'] = info_base['Case Type:']

    out_org_body(sncp[org]['SheetName'], org)




# ===================================================
if __name__ == '__main__':


    # Create new excel file
    new_wb = openpyxl.Workbook()
    init_style()
    list_col = gen_list_col()





    org_path = glob.glob(src_path)
    for i in range(len(org_path)):
        org_name = org_path[i].split('\\')[-1]
        if org_name not in sncp:
            sncp[org_name] = {'SheetName': str(i + 1), 'CurRow': 1}

    for key in reversed(list(sncp.keys())):
        new_wb.create_sheet(title=sncp[key]['SheetName'], index=0)
        if key not in sncp_lock:
            make_org_topic(sncp[key]['SheetName'], key)

    #print(sncp)
    make_list()  # Список листов с гиперссылкой
    make_raw_topic()

    for i in range(len(org_path)):
        print('----------------------------------------------------------------------------')
        print(org_path[i])
        workplaces = glob.glob(org_path[i]+'\*')
        hwi_htm_files = glob.glob(org_path[i]+'\*\*.htm*')

        org_name = org_path[i].split('\\')[-1]
        j = 0
        for src_tmp_path in hwi_htm_files:

            print(org_name)
            scan_hwi_htm(src_tmp_path)
            j += 1
            if j > 7:
                break
        new_wb.save(dst_path + dst_file)
        if i > 5:
            break


















# ========================================

    new_wb.save(dst_path + dst_file)
    print('Save file')



